import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import time
from PIL import Image, ImageDraw, ImageFont, ImageOps
import os
import io
import re
import google.generativeai as genai
from datetime import datetime
import urllib.parse 
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import requests 
from bs4 import BeautifulSoup 
import traceback 

# Carica tutte le variabili dal file .env (deve essere la prima cosa)
load_dotenv()

# --- CONFIGURAZIONE ---
# ⚠️ 1. TOKEN TELEGRAM
API_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN') 
# ⚠️ 2. CHIAVE GOOGLE (Usata SOLO per recupero emoji)
GOOGLE_API_KEY = os.getenv('GOOGLE_API_KEY') 
# ⚠️ 3. CANALE
CHANNEL_ID = '@citazioneradar' 
# ⚠️ 4. CHAT ID PERSONALE per notifiche critiche
FABRIZIO_CHAT_ID = os.getenv('FABRIZIO_CHAT_ID') 

AMAZON_TAG = 'radartest-21' 
DB_FILE = "Registro_Vendite.xlsx"
FONT_NAME = "Montserrat-Bold.ttf"

# Link Disclaimer
LINK_INFO_POST = "https://t.me/citazioneradar/178" 

# --- VARIABILI DI STATO GLOBALI ---
# RIMOZIONE DI 'is_posting_open' - Non più necessaria.

# --- POST GIORNALIERI (BUONGIORNO / BUONANOTTE) ---
POST_BUONGIORNO = (
    "☀️ **BUONGIORNO & BUON INIZIO SETTIMANA!** ☕\n\n"
    "Il ProfitBot è attivo e pronto a scovare le migliori offerte del giorno.\n"
    "Restate connessi sul canale per non perdere i ribassi più importanti!"
)

POST_BUONANOTTE = (
    "🌙 **Buonanotte!** ✨\n\n"
    "La caccia alle offerte per oggi è terminata.\n"
    "Domattina si ricomincia, vi aspettiamo!"
)
# --- FINE POST GIORNALIERI ---


# --- CONFIGURAZIONE IA ---
try:
    if GOOGLE_API_KEY:
        genai.configure(api_key=GOOGLE_API_KEY)
        model = genai.GenerativeModel('gemini-1.5-flash')
    else:
        print("⚠️ Chiave GOOGLE_API_KEY mancante. Le emoji IA non funzioneranno.")
        model = None
except Exception as e:
    print(f"Errore configurazione Gemini: {e}")
    model = None

# Verifica TOKEN
if not API_TOKEN:
    print("❌ ERRORE CRITICO: TELEGRAM_BOT_TOKEN non trovato nel file .env.")
    pass 

bot = telebot.TeleBot(API_TOKEN)
user_data = {} 

print("✅ ProfitBot Definitivo AVVIATO!")

# --- GESTIONE ERRORI CRITICI ---
def handle_critical_error(e):
    error_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    error_traceback = traceback.format_exc()
    message = (
        f"🚨 **ERRORE CRITICO in ProfitBot!** 🚨\n"
        f"⏰ Data/Ora: `{error_time}`\n"
        f"🐞 Tipo Errore: `{type(e).__name__}`\n"
        f"💬 Messaggio: `{str(e)}`\n\n"
        f"--- TRACEBACK ---\n"
        f"```\n"
        f"{error_traceback[:1000]}..." 
        f"```"
    )
    # Controllo che FABRIZIO_CHAT_ID sia impostato nel .env e valido
    if FABRIZIO_CHAT_ID:
        try:
            bot.send_message(FABRIZIO_CHAT_ID, message, parse_mode='Markdown')
        except Exception as notify_e:
            print(f"Impossibile inviare la notifica a {FABRIZIO_CHAT_ID}: {notify_e}")
    print(message) 

# --- DATABASE (Invariato nella logica) ---
def inizializza_db():
    if not os.path.exists(DB_FILE):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Vendite"
            ws.append(["DATA", "ORA", "ASIN", "TITOLO", "PREZZO VECCHIO", "PREZZO NUOVO", "LINK", "FILE_ID"]) 
            ws.column_dimensions['D'].width = 50
            wb.save(DB_FILE)
        except Exception as e: 
            print(f"Errore inizializzazione DB: {e}")

def is_gia_pubblicato(asin):
    if not os.path.exists(DB_FILE): return False
    try:
        wb = load_workbook(DB_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and asin in str(row[2]) and asin != "NO_ASIN":
                return True
    except:
        return False
    return False

def salva_in_excel(dati):
    adesso = datetime.now()
    try:
        if os.path.exists(DB_FILE):
            wb = load_workbook(DB_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
        
        ws.append([
            adesso.strftime("%d/%m/%Y"), 
            adesso.strftime("%H:%M:%S"), 
            dati.get('asin', 'N/A'), 
            dati.get('titolo', 'Nessun Titolo'), 
            dati.get('old_fmt_save', '0,00'),   
            dati.get('new_fmt_save', '0,00'),   
            dati.get('link', ''),
            dati.get('file_id', 'N/A') 
        ])
        wb.save(DB_FILE)
        return True
    except Exception as e:
        print(f"Errore salvataggio Excel: {e}")
        return False

# --- CLEANER E UTILITY (Invariato) ---
def clean_input_price(text):
    text = text.replace('€', '').replace('EUR', '').strip()
    return text

def clean_price_calc(price_str):
    if not isinstance(price_str, str):
        return 0.0
    price_str = price_str.replace('€', '').replace('$', '').strip()
    if price_str.count(',') == 1 and price_str.count('.') <= 1 and price_str.rfind(',') > price_str.rfind('.'):
        price_str = price_str.replace('.', '')
        price_str = price_str.replace(',', '.')
    else:
        price_str = price_str.replace(',', '') 
        
    try:
        return float(price_str)
    except ValueError:
        return 0.0 
        
def format_price_euro(val_float):
    return "{:,.2f}".format(val_float).replace(",", "X").replace(".", ",").replace("X", ".")

def format_price_for_excel(val_float):
    return "{:.2f}".format(val_float).replace('.', ',')

def sanitize_description(text):
    if not text: return text
    pattern = r'(\d+[\.,]?\d*["\w\s]*)(\1)'
    cleaned_text = re.sub(pattern, r'\1', text, flags=re.IGNORECASE)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
    return cleaned_text

def escape_markdown(text):
    """
    Rende la stringa sicura per il parsing MarkdownV2 di Telegram.
    Sostituisce i caratteri speciali con la loro versione 'escaped'.
    """
    text = text.replace('\\', r'\\') # Deve essere il primo!
    for char in ['*', '_', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']:
        text = text.replace(char, r'\\' + char)
    return text

# --- IA E EMOJI (Invariato) ---
def rewrite_with_ai(testo_grezzo):
    if not testo_grezzo:
        return ""
    testo = testo_grezzo.strip().replace('"', '').replace('**', '')
    if testo.endswith('...'):
        testo = testo[:-3].strip()
    return testo
        
def get_emoji_from_ia(titolo):
    if not titolo: return "📦"
    
    titolo_lower = titolo.lower().replace('.', '').replace(',', '')
    
    emoji_map = {
        "friggitrice": "🍳", "forno": "🍗", "frullatore": "🍹", "macchina da caffè": "☕",
        "pentola": "🍲", "padella": "🍳", "aspirapolvere": "🧹", "robot aspirapolvere": "🤖",
        "detersivo": "🧺", "ammorbidente": "🧺", "finish": "🍽️", 
        "smartphone": "📱", "tablet": "📱", "notebook": "💻", 
        "cuffie": "🎧", "auricolari": "🎧", "solo": "🎧",
        "custodia per": "📱", "t-shirt": "👕", "jeans": "👖", 
        "scarpe": "👟", "spazzola ad aria": "💆‍♀️", "prosecco": "🍾", 
        "vino": "🍷", "cioccolato": "🍫", "giacca": "🧥", 
        "piumino": "🧥", "tuta": "👕", "chiavetta": "💾", 
        "spazzolino": "🦷", 
        "elettrico": "⚡",   
        "bambino": "🧸",    
        "bebes": "🧸",      
        "chiccò": "🧸",     
        "i-master": "🍹",   
    }
    
    for keyword, emoji in emoji_map.items():
        if keyword in titolo_lower:
            return emoji 

    if not model: return "📦"
    
    prompt = (
        f"Analizza il titolo: '{titolo}'. Restituisci SOLO UNA singola emoji molto specifica "
        f"che rappresenti il prodotto. Non usare emoji generiche come 📦, 🎁. Se non trovi una corrispondenza, restituisci solo 📦."
    )
    
    try:
        response = model.generate_content(prompt)
        raw_text = response.text.strip()
        
        emoji_match = re.search(r'([\U0001F000-\U001FFFFF])', raw_text)
        
        if emoji_match:
            return emoji_match.group(1) 
        
        return "📦"
        
    except Exception as e:
        print(f"Errore IA nel recupero emoji: {e}")
        return "📦" 

# --- FUNZIONE: WEB SCRAPING ---
def get_product_data(url):
    """Estrae titolo e prezzo corrente da un link Amazon."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
        'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 503 or "captcha" in response.text.lower():
             print("❌ Scraping fallito: CAPTCHA o blocco di Amazon (Status 503).")
             return None

        soup = BeautifulSoup(response.content, 'html.parser')
        
        # 1. ESTREZIONE TITOLO
        title_tag = soup.find('span', id='productTitle')
        title = title_tag.text.strip() if title_tag else None

        # 2. ESTREZIONE PREZZO (Targeting multipli ID/classi per robustezza)
        price = None
        
        price_tag_main = soup.find('span', class_='a-price aok-align-center') 
        if price_tag_main:
            price_text = price_tag_main.find('span', class_='a-offscreen')
            if price_text:
                price = price_text.text.strip()
        
        if not price:
            price_tag_lampo = soup.find('span', id='priceblock_ourprice') or soup.find('span', id='priceblock_dealprice')
            if price_tag_lampo:
                price = price_tag_lampo.text.strip()

        # Pulizia e conversione prezzo
        cleaned_price = 0.0
        if price:
            cleaned_price = clean_price_calc(price)
            if cleaned_price == 0.0:
                price = None 
                
        # 3. RITORNO DATI
        if title and cleaned_price > 0.0:
            return {
                'titolo_estratto': rewrite_with_ai(title), 
                'prezzo_nuovo_estratto': cleaned_price 
            }
        
        print(f"⚠️ Scraping riuscito, ma dati incompleti. Titolo: {title}, Prezzo pulito: {cleaned_price}")
        return None

    except requests.exceptions.Timeout:
        print("❌ Scraping fallito: Timeout.")
        return None
    except Exception as e:
        print(f"❌ Scraping fallito per errore generico: {e}")
        return None
# --- FINE FUNZIONE WEB SCRAPING ---

# --- FUNZIONI DI RIASSUNTO E COLLAGE ---
def get_riassunto_offerte():
    if not os.path.exists(DB_FILE):
        return "⚠️ Nessun record di offerte trovate nel database."

    adesso = datetime.now()
    oggi_str = adesso.strftime("%d/%m/%Y")
    riepilogo = f"👀 *Ecco il riassunto delle migliori offerte di oggi!* \n\n" 
    
    record_oggi = {}
    
    try:
        # ... (omissis, la logica di caricamento DB è invariata)

        wb = load_workbook(DB_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_db = row[0]
            prezzo_vecchio_str = row[4] 
            prezzo_nuovo_str = row[5]

            old_val = clean_price_calc(prezzo_vecchio_str)
            new_val = clean_price_calc(prezzo_nuovo_str)

            if str(data_db) == oggi_str and row[2] != 'N/A' and old_val > new_val and old_val > 0:
                perc = int(100 - (new_val / old_val * 100))
                old_fmt = format_price_euro(old_val) + "€"
                new_fmt = format_price_euro(new_val) + "€"
                
                record_oggi[row[2]] = {
                    'titolo': row[3],
                    'prezzo_vecchio_fmt': old_fmt,
                    'prezzo_nuovo_fmt': new_fmt,
                    'link': row[6],
                    'sconto_perc': perc 
                }

        num_offerte = len(record_oggi)
        if num_offerte == 0:
            return "⚠️ Nessuna offerta di ribasso trovata oggi nel database."
            
        offerte_ordinate = sorted(record_oggi.items(), key=lambda item: item[1]['sconto_perc'], reverse=True)
            
        count = 0 
        for asin, dati in offerte_ordinate:
            if count >= 5: break
            titolo_pulito = dati['titolo'].split('(')[0].strip()
            
            # Applichiamo escape_markdown SOLO al titolo, non al link.
            titolo_pulito_sicuro = escape_markdown(titolo_pulito)
            
            riepilogo += f"🚨 **{titolo_pulito_sicuro}**\n" 
            riepilogo += f"💰 **{dati['prezzo_nuovo_fmt']}** _invece di {dati['prezzo_vecchio_fmt']}_\n"
            
            short_link_base = dati['link'].split('?')[0] 
            
            asin_match = re.search(r'(B0[A-Z0-9]{8})', short_link_base)
            if asin_match:
                short_link = f"https://www.amazon.it/dp/{asin_match.group(1)}?tag={AMAZON_TAG}"
            else:
                short_link = dati['link'] 

            # 💡 CORREZIONE LINK: Usiamo la sintassi [Testo](URL) in Markdown.
            # L'URL (short_link) NON DEVE essere escaped.
            riepilogo += f"🔍 [Link Prodotto]({short_link})\n\n" 
            count += 1
            
        riepilogo += "🔗 *Tutti i link sono affiliati.*\n"
        
        # ⚠️ CORREZIONE LUNGHEZZA CAPTION
        MAX_CAPTION_LENGTH = 980 
        if len(riepilogo) > MAX_CAPTION_LENGTH:
             riepilogo = riepilogo[:MAX_CAPTION_LENGTH] + "\n\n[...] *Riepilogo Troncato per limite Telegram.*"
             
        return riepilogo
        
    except Exception as e:
        return f"❌ Errore nella lettura DB: {e}"

def get_latest_image_ids():
    if not os.path.exists(DB_FILE): return []
    file_ids = []
    adesso = datetime.now()
    oggi_str = adesso.strftime("%d/%m/%Y")
    ws = None
    
    try:
        wb = load_workbook(DB_FILE)
        ws = wb.active
        if ws.max_row < 2: return []
            
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            data_db = row[0]
            file_id = row[7] 
            prezzo_vecchio_str = row[4]
            prezzo_nuovo_str = row[5]

            old_val = clean_price_calc(prezzo_vecchio_str)
            new_val = clean_price_calc(prezzo_nuovo_str)
            
            if (str(data_db) == oggi_str and 
                file_id != 'N/A' and 
                file_id not in file_ids and 
                old_val > new_val):

                file_ids.append(file_id)
                if len(file_ids) >= 5: break
        
        return file_ids
    except Exception as e:
        print(f"Errore lettura file IDs: {e}") 
        return []

def crea_collage_riassunto():
    file_ids = get_latest_image_ids()
    
    if not file_ids:
        img = Image.new('RGB', (600, 600), color = 'gray')
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype(FONT_NAME, 40)
        except:
            font = ImageFont.load_default()
            
        draw.text((50, 280), "⚠️ NESSUNA FOTO DI RIBASSO TROVATA OGGI ⚠️", fill=(255, 255, 255), font=font)
        bio = io.BytesIO()
        img.save(bio, 'JPEG', quality=90)
        bio.seek(0)
        return bio
        
    immagini_collage = []
    
    for file_id in file_ids:
        try:
            file_info = bot.get_file(file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            img = Image.open(io.BytesIO(downloaded_file)).convert("RGB")
            immagini_collage.append(img)
        except Exception as e:
            print(f"Errore download foto {file_id}: {e}")
            continue

    if not immagini_collage:
        img = Image.new('RGB', (600, 600), color = 'gray')
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype(FONT_NAME, 40)
        except:
            font = ImageFont.load_default()
            
        draw.text((50, 280), "⚠️ FALLBACK: ERRORE SCARICAMENTO FOTO ⚠️", fill=(255, 255, 255), font=font)
        bio = io.BytesIO()
        img.save(bio, 'JPEG', quality=90)
        bio.seek(0)
        return bio

    num_immagini = len(immagini_collage)
    
    COLLAGE_BASE_DIM = 800
    PADDING = 10 
    
    if num_immagini <= 2:
        cols = num_immagini
        rows_effettive = 1
    elif num_immagini == 3:
        cols = 3
        rows_effettive = 1
    elif num_immagini <= 5: 
        cols = 3 
        rows_effettive = 2 
    else: 
        cols = 3
        rows_effettive = 3 
    
    dim_cella_effettiva = COLLAGE_BASE_DIM // cols
    larghezza_finale = cols * dim_cella_effettiva
    altezza_finale = rows_effettive * dim_cella_effettiva
    
    collage = Image.new('RGB', (larghezza_finale, altezza_finale), color='white')
    
    for i, img in enumerate(immagini_collage):
        
        row = i // cols
        col = i % cols
        
        img_resized = ImageOps.fit(img, (dim_cella_effettiva - PADDING, dim_cella_effettiva - PADDING), Image.Resampling.LANCZOS)
        x_offset = col * dim_cella_effettiva + PADDING // 2
        y_offset = row * dim_cella_effettiva + PADDING // 2
        
        if num_immagini == 4 or num_immagini == 5:
            if i < 2:
                x_offset = col * dim_cella_effettiva + dim_cella_effettiva // 2 + PADDING // 2 
                y_offset = PADDING // 2 
            else:
                 col_riga2 = i - 2 
                 x_offset = col_riga2 * dim_cella_effettiva + PADDING // 2
                 y_offset = dim_cella_effettiva + PADDING // 2 
        
        collage.paste(img_resized, (x_offset, y_offset))

    bio = io.BytesIO()
    collage.save(bio, 'JPEG', quality=95)
    bio.seek(0)
    return bio

# --- GESTIONE FLUSSO PUBBLICAZIONI RIASSUNTO E STATO GLOBALE ---

# Rimosse open_posts_handler e close_posts_handler

def show_riassunto(chat_id, call):
    """Genera il riassunto testuale e chiede conferma per il collage."""
    riepilogo_txt = get_riassunto_offerte()
    
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("📸 Crea Collage e Pubblica", callback_data="pubblica_riassunto"))
    markup.add(InlineKeyboardButton("❌ Annulla", callback_data="reset_all"))
    
    bot.edit_message_text(
        riepilogo_txt, 
        chat_id, 
        call.message.message_id, 
        reply_markup=markup, 
        parse_mode='Markdown'
    )

def pubblica_riassunto_handler(chat_id, call):
    """Invia il collage al canale di test (chat_id) e chiede la conferma di pubblicazione."""
    try:
        collage_bytes = crea_collage_riassunto()
        riepilogo_txt = get_riassunto_offerte() 
        
        markup_confirm = InlineKeyboardMarkup(row_width=1)
        markup_confirm.add(InlineKeyboardButton("✅ CONFERMA PUBBLICAZIONE", callback_data="confirma_pubblica_riassunto"))
        markup_confirm.add(InlineKeyboardButton("❌ Annulla", callback_data="reset_all"))
        
        # 1. Invio la foto e il riepilogo nel canale di TEST (chat_id)
        msg_photo = bot.send_photo(chat_id, collage_bytes, caption=riepilogo_txt, parse_mode='Markdown')
        
        # 2. Chiedo conferma, memorizzando l'ID del messaggio della foto e della conferma
        msg_confirm = bot.send_message(
            chat_id, 
            "**Anteprima creata e inviata qui sopra.**\n\n"
            "Confermi l'invio sul canale pubblico? Dopo la pubblicazione, l'anteprima verrà *cancellata*.",
            reply_markup=markup_confirm,
            parse_mode='Markdown'
        )
        
        # Memorizzo gli ID per la conferma e cancellazione successive
        user_data[chat_id]['riassunto_photo_id'] = msg_photo.message_id
        user_data[chat_id]['riassunto_confirm_id'] = msg_confirm.message_id

        # Cancello il pulsante 'Crea Collage' per pulizia
        bot.delete_message(chat_id, call.message.message_id)
        
    except Exception as e:
        bot.send_message(chat_id, f"❌ ERRORE: Impossibile creare o inviare il riassunto: {e}")
        handle_critical_error(e)


def confirma_pubblica_riassunto(chat_id, call):
    """
    Pubblica effettivamente il riassunto sul canale pubblico e cancella l'anteprima.
    (Modificato per risolvere l'errore 400 'no text in the message to edit')
    """
    
    if chat_id not in user_data or 'riassunto_photo_id' not in user_data[chat_id]:
        bot.send_message(chat_id, "⚠️ Sessione riassunto scaduta. Riprova da Menu Principale.", reply_markup=main_menu())
        return

    try:
        # Prendo i dati
        photo_id_to_copy = user_data[chat_id]['riassunto_photo_id']
        confirm_msg_id = user_data[chat_id]['riassunto_confirm_id']
        
        # 1. Copio il messaggio della foto dal chat_id al CHANNEL_ID
        bot.copy_message(
            chat_id=CHANNEL_ID, 
            from_chat_id=chat_id, 
            message_id=photo_id_to_copy
        )
        
        # 2. Aggiorno il messaggio di conferma (il testo che contiene il bottone)
        bot.edit_message_text(
            "✅ **PUBBLICATO SUL CANALE!**", 
            chat_id, 
            confirm_msg_id, 
            reply_markup=None, 
            parse_mode='Markdown'
        )
        
        # 3. Cancello l'anteprima della foto dal chat_id
        bot.delete_message(chat_id, photo_id_to_copy)
        
        # 4. Cleanup
        del user_data[chat_id]['riassunto_photo_id']
        del user_data[chat_id]['riassunto_confirm_id']
        
        bot.send_message(chat_id, "Ottimo lavoro! Procediamo?", reply_markup=main_menu())
        
    except Exception as e:
        # In caso di errore, edita il messaggio di conferma per notificare
        error_msg = f"❌ ERRORE PUBBLICAZIONE: {e}"
        try:
             # Tentiamo di modificare il messaggio di conferma esistente
             bot.edit_message_text(
                 error_msg, 
                 chat_id, 
                 call.message.message_id, 
                 reply_markup=main_menu()
             )
        except Exception as edit_e:
             # Fallback se non riusciamo a modificare (es. se l'errore è avvenuto troppo tardi)
             bot.send_message(chat_id, error_msg, reply_markup=main_menu())
             print(f"Errore secondario editing: {edit_e}")

        handle_critical_error(e)

# --- NUOVA FUNZIONE PER I POST GIORNALIERI ---

def send_daily_post(chat_id, call, post_content, post_name):
    """Invia un messaggio predefinito (Buongiorno/Buonanotte) al canale."""
    try:
        # Invio il post al canale pubblico
        bot.send_message(CHANNEL_ID, post_content, parse_mode='Markdown')

        # Aggiorno l'interfaccia nel bot per feedback
        bot.edit_message_text(
            f"✅ **'{post_name}' pubblicato con successo sul canale!**", 
            chat_id, 
            call.message.message_id, 
            reply_markup=main_menu(), 
            parse_mode='Markdown'
        )
    except Exception as e:
        error_msg = f"❌ ERRORE: Impossibile pubblicare '{post_name}' sul canale. Controlla il CHANNEL_ID o i permessi: {e}"
        bot.send_message(chat_id, error_msg)
        handle_critical_error(e)
        # Riporto il menu principale
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=main_menu())


# --- MENU e GESTORE CALLBACK ---
def main_menu():
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("🚀 Inizia Nuovo Post", callback_data="url_libero"))
    markup.add(InlineKeyboardButton("🗒️ Riassunto Offerte Oggi", callback_data="show_riassunto"))
    
    # Pulsanti Buongiorno/Buonanotte
    markup.add(InlineKeyboardButton("☀️ Pubblica Buongiorno", callback_data="pubblica_buongiorno"), 
               InlineKeyboardButton("🌙 Pubblica Buonanotte", callback_data="pubblica_buonanotte"))
               
    # RIMOSSI i pulsanti "Inizio Pubblicazioni per Oggi" e "Chiudi Pubblicazioni di Oggi"
    
    markup.add(InlineKeyboardButton("🏠 Menu Principale (Start)", callback_data="reset_all")) 
    return markup

def get_extra_markup(chat_id):
    extras = user_data.get(chat_id, {}).get('extras', {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False})
    
    txt_prime = "✅ Prime" if extras['prime'] else "🚚 Prime?"
    txt_lampo = "✅ Lampo" if extras['lampo'] else "⚡️ Offerta a tempo?"
    txt_choice = "✅ Scelta Amazon" if extras['choice'] else "⭐ Scelta Amazon?"
    coupon_val = extras.get('coupon')
    txt_coupon = f"✅ {coupon_val}" if coupon_val else "🎫 Coupon?"
    txt_rapida = "✅ Vendita Rapida" if extras['rapida'] else "🔥 Offerta a vendita rapida?" 
    
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(InlineKeyboardButton(txt_prime, callback_data="toggle_prime"),
               InlineKeyboardButton(txt_lampo, callback_data="toggle_lampo"))
    markup.add(InlineKeyboardButton(txt_choice, callback_data="toggle_choice"),
               InlineKeyboardButton(txt_coupon, callback_data="set_coupon"))
    markup.add(InlineKeyboardButton(txt_rapida, callback_data="toggle_rapida"))
    markup.add(InlineKeyboardButton("📸 CONTINUA", callback_data="finish_extras"))
    return markup

@bot.message_handler(commands=['start'])
def welcome(message):
    bot.clear_step_handler_by_chat_id(message.chat.id)
    user_data[message.chat.id] = {'extras': {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False}} 
    bot.send_message(message.chat.id, "🚨 *ProfitBot di Radar Offerte*\nFlusso pronto per l'automazione.", reply_markup=main_menu(), parse_mode='Markdown')

# --- GESTORE CALLBACK (LOGICA FLUSSO CRITICA) ---
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    chat_id = call.message.chat.id
    try: bot.answer_callback_query(call.id)
    except: pass

    RESET_EXTRAS = {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False}

    if call.data == "reset_all":
        bot.clear_step_handler_by_chat_id(chat_id) 
        user_data[chat_id] = {'extras': RESET_EXTRAS} 
        bot.edit_message_text("⭐️ *ProfitBot Stabile*\nFlusso pronto per l'automazione.", chat_id, call.message.message_id, reply_markup=main_menu(), parse_mode='Markdown')
        return
        
    if call.data == "force_unlock" or call.data == "url_libero":
        bot.clear_step_handler_by_chat_id(chat_id) 
        user_data[chat_id] = {'extras': RESET_EXTRAS} 
        ask_link(chat_id)
        return

    # Gestione post giornalieri
    if call.data == "pubblica_buongiorno": 
        send_daily_post(chat_id, call, POST_BUONGIORNO, "Buongiorno")
        return
    elif call.data == "pubblica_buonanotte": 
        send_daily_post(chat_id, call, POST_BUONANOTTE, "Buonanotte")
        return
        
    # Gestione riassunto
    elif call.data == "show_riassunto": show_riassunto(chat_id, call); return
    elif call.data == "pubblica_riassunto": pubblica_riassunto_handler(chat_id, call); return
    elif call.data == "confirma_pubblica_riassunto": confirma_pubblica_riassunto(chat_id, call); return
    
    # RIMOZIONE della gestione per "close_posts" e "open_posts"

    # NAVIGAZIONE E CORREZIONE MANUALE (Invariato)
    elif call.data == "back_to_link": ask_link(chat_id)
    # Nel flusso automatico, i back to title/old/new sono meno usati, ma li manteniamo per il fallback
    elif call.data == "back_to_title": ask_title(chat_id) # CORREZIONE 1: ask_old_price -> ask_title
    elif call.data == "back_to_old": ask_old_price_manuale(chat_id)
    elif call.data == "back_to_new": ask_new_price_manuale(chat_id) # Corretto nome funzione
    elif call.data == "back_to_reviews": ask_reviews(chat_id)
    elif call.data == "back_to_description": ask_description(chat_id)
    
    # NUOVE FUNZIONI DI CORREZIONE AUTOMATICA
    elif call.data == "correct_title_auto":
        msg = bot.send_message(chat_id, "✍️ **Inserire il Titolo Corretto:**", reply_markup=get_nav_markup("back_to_link"), parse_mode='Markdown')
        bot.register_next_step_handler(msg, step_correct_title)
        
    elif call.data == "correct_new_price_auto":
        msg = bot.send_message(chat_id, "✍️ **Inserire il NUOVO Prezzo Corretto:**", reply_markup=get_nav_markup("back_to_link"), parse_mode='Markdown')
        bot.register_next_step_handler(msg, step_correct_new_price)
        
    # GESTIONE EXTRA
    elif call.data == "finish_description_step":
        step_ask_extras(chat_id)
        return

    if 'extras' not in user_data.get(chat_id, {}): 
        user_data[chat_id]['extras'] = RESET_EXTRAS

    elif call.data == "toggle_prime":
        user_data[chat_id]['extras']['prime'] = not user_data[chat_id]['extras']['prime']
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=get_extra_markup(chat_id))
    elif call.data == "toggle_lampo":
        user_data[chat_id]['extras']['lampo'] = not user_data[chat_id]['extras']['lampo']
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=get_extra_markup(chat_id))
    elif call.data == "toggle_choice": 
        user_data[chat_id]['extras']['choice'] = not user_data[chat_id]['extras']['choice']
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=get_extra_markup(chat_id))
    elif call.data == "toggle_rapida": 
        user_data[chat_id]['extras']['rapida'] = not user_data[chat_id]['extras']['rapida']
        bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=get_extra_markup(chat_id))
        
    elif call.data == "set_coupon":
        msg = bot.send_message(chat_id, "🎫 **Valore Coupon:**", parse_mode='Markdown')
        bot.register_next_step_handler(msg, step_coupon_input)
    elif call.data == "finish_extras":
        ask_photo(chat_id)

    # PUBBLICAZIONE SINGOLA (Invariato)
    elif call.data == "pubblica_si":
        if chat_id in user_data and 'img_bytes' in user_data[chat_id]:
            dati = user_data[chat_id]
            try:
                photo_file = io.BytesIO(dati['img_bytes'])
                bot.send_photo(CHANNEL_ID, photo_file, caption=dati['caption'], reply_markup=dati['markup'], parse_mode='Markdown')
                salva_in_excel(dati)
                
                msg = "✅ **PUBBLICATO!**"
                # Modifichiamo il messaggio di conferma (il testo)
                bot.edit_message_text(msg, chat_id, call.message.message_id, parse_mode='Markdown')

                bot.send_message(chat_id, "Procediamo?", reply_markup=main_menu()) 
                user_data[chat_id] = {'extras': RESET_EXTRAS} 
            except Exception as e:
                bot.send_message(chat_id, f"❌ ERRORE TECNICO: {e}", parse_mode='Markdown')
                handle_critical_error(e)
        else:
            bot.send_message(chat_id, "⚠️ Sessione scaduta.", parse_mode='Markdown')

    elif call.data == "pubblica_no":
        bot.edit_message_text("❌ Annullato.", chat_id, call.message.message_id, parse_mode='Markdown')
        bot.send_message(chat_id, "Ok, riproviamo.", reply_markup=main_menu(), parse_mode='Markdown')

# --- FUNZIONI STEP (Invariato) ---

def get_nav_markup(step_back=None):
    markup = InlineKeyboardMarkup()
    if step_back:
        if step_back == "back_to_link":
            markup.add(InlineKeyboardButton("🔙 Correggi Link", callback_data=step_back))
        elif step_back == "back_to_title":
            markup.add(InlineKeyboardButton("🔙 Correggi Titolo", callback_data=step_back))
        elif step_back == "back_to_old":
            markup.add(InlineKeyboardButton("🔙 Correggi Vecchio Prezzo", callback_data=step_back))
        elif step_back == "back_to_new":
            markup.add(InlineKeyboardButton("🔙 Correggi Nuovo Prezzo", callback_data=step_back))
        elif step_back == "back_to_reviews":
            markup.add(InlineKeyboardButton("🔙 Correggi Voto/Recensioni", callback_data=step_back))
        elif step_back == "back_to_description":
            markup.add(InlineKeyboardButton("🔙 Correggi Descrizione", callback_data=step_back)) 

    markup.add(InlineKeyboardButton("❌ ANNULLA", callback_data="reset_all"))
    return markup

def ask_link(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "1️⃣ **Incolla il Link Amazon:**", reply_markup=get_nav_markup(None), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_link) 

# Flusso Manuale di Fallback (come prima del codice scraping)
def ask_title(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "2️⃣ **Inserire il Titolo del Prodotto:**", reply_markup=get_nav_markup("back_to_link"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_titolo_manuale) # Usiamo la funzione manuale

def step_titolo_manuale(message):
    if message.text and message.text.startswith('/'): return
    chat_id = message.chat.id
    titolo = rewrite_with_ai(message.text) 
    user_data[chat_id]['titolo'] = titolo
    bot.send_message(chat_id, f"✨ {titolo}") 
    ask_old_price_manuale(chat_id)

def ask_old_price_manuale(chat_id): # Flusso Manuale di Fallback: chiede il vecchio prezzo
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "3️⃣ **Prezzo Vecchio (es: 1499.00):**", reply_markup=get_nav_markup("back_to_link"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_prezzo_old_manuale)

def step_prezzo_old_manuale(message):
    if message.text and message.text.startswith('/'): return
    pulito = clean_input_price(message.text)
    val = clean_price_calc(pulito) 
    if val == 0.0 and pulito.strip() != "0" and pulito.strip() != "0,00":
        bot.send_message(message.chat.id, "❌ Prezzo Vecchio non valido. Riprova.")
        ask_old_price_manuale(message.chat.id)
        return

    user_data[message.chat.id]['old'] = val
    ask_new_price_manuale(message.chat.id)

def ask_new_price_manuale(chat_id): # Flusso Manuale di Fallback: chiede il nuovo prezzo
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "4️⃣ **Prezzo NUOVO (es: 1019.99):**", reply_markup=get_nav_markup("back_to_old"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_prezzo_new_manuale)

def step_prezzo_new_manuale(message):
    # Riprende la logica originale di step_prezzo_new
    if message.text and message.text.startswith('/'): return
    pulito = clean_input_price(message.text)
    val = clean_price_calc(pulito) 
    
    if val == 0.0 and pulito.strip() != "0" and pulito.strip() != "0,00":
        bot.send_message(message.chat.id, "❌ Prezzo Nuovo non valido. Riprova.")
        ask_new_price_manuale(message.chat.id)
        return
    
    chat_id = message.chat.id
    user_data[chat_id]['new'] = val
    
    # Calcola sconto
    step_update_discount(chat_id)
    
    ask_reviews(chat_id)

# Funzioni Reviews, Descrizione, Foto (Invariate)
def ask_reviews(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "⭐ **Voto e Recensioni:**\nEs: `284 4.5`\n(Scrivi 'no' per saltare)", reply_markup=get_nav_markup("back_to_new"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_reviews)

def step_reviews(message):
    if message.text and message.text.startswith('/'): return
    txt = message.text.strip()
    if txt.lower() == "no":
        user_data[message.chat.id]['voto'] = None
    else:
        parti = txt.split()
        if len(parti) >= 2:
            num = parti[0]
            voto = parti[1]
            try: num = "{:,}".format(int(num.replace('.', '').replace(',', ''))).replace(',', '.')
            except: pass
            user_data[message.chat.id]['voto'] = f"⭐️ {num} Recensioni: {voto} / 5.0"
        else:
            user_data[message.chat.id]['voto'] = f"⭐️ {txt}"
            
    ask_description(message.chat.id) 
    
def ask_description(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "5️⃣ **Descrizione Prodotto/Marketing**:\n(Max 4 righe brevi. Scrivi 'no' per saltare)", reply_markup=get_nav_markup("back_to_reviews"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_description)
    
def step_description(message):
    if message.text and message.text.startswith('/'): return
    chat_id = message.chat.id
    txt_grezzo = message.text.strip()
    
    if txt_grezzo.lower() == "no":
        user_data[chat_id]['descrizione'] = None
        messaggio_anteprima = "✅ Descrizione saltata. Clicca Continua per le opzioni extra."
    else:
        sanitized_txt = sanitize_description(txt_grezzo) 
        user_data[chat_id]['descrizione'] = sanitized_txt
        messaggio_anteprima = f"✅ **Descrizione registrata:**\n`{sanitized_txt}`"

    markup_confirm = InlineKeyboardMarkup(row_width=2)
    markup_confirm.add(InlineKeyboardButton("✅ CONTINUA (Opzioni Extra)", callback_data="finish_description_step"))
    markup_confirm.add(InlineKeyboardButton("✍️ Modifica Descrizione", callback_data="back_to_description")) 
    markup_confirm.add(InlineKeyboardButton("❌ ANNULLA POST", callback_data="reset_all")) 

    bot.send_message(chat_id, messaggio_anteprima, parse_mode='Markdown')
    bot.send_message(chat_id, "Se la descrizione è corretta, prosegui:", reply_markup=markup_confirm)

def step_ask_extras(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "🚨 **Opzioni Extra (o Invia subito la FOTO):**", reply_markup=get_extra_markup(chat_id), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_check_extras_photo)

def step_check_extras_photo(message):
    if message.content_type == 'photo':
        step_foto_process(message)
    else:
        msg = bot.send_message(message.chat.id, "❌ Per favore, invia la FOTO oppure clicca i pulsanti (es: CONTINUA).")
        bot.register_next_step_handler(msg, step_check_extras_photo)

def step_coupon_input(message):
    chat_id = message.chat.id
    valore = message.text
    if user_data.get(chat_id) is None: 
        user_data[chat_id] = {'extras': {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False}}
    if 'extras' not in user_data[chat_id]: 
        user_data[chat_id]['extras'] = {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False}
        
    user_data[chat_id]['extras']['coupon'] = valore
    msg = bot.send_message(chat_id, f"✅ Coupon '{valore}' salvato! Ora FOTO:", reply_markup=get_extra_markup(chat_id))
    bot.register_next_step_handler(msg, step_check_extras_photo)

def ask_photo(chat_id):
    bot.clear_step_handler_by_chat_id(chat_id)
    msg = bot.send_message(chat_id, "6️⃣ **Invia FOTO:**", reply_markup=get_nav_markup("back_to_description"), parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_foto_process)

def step_foto_process(message):
    if message.text and message.text.startswith('/'): return
    if not message.photo: 
        msg = bot.send_message(message.chat.id, "❌ Devi inviare la FOTO del prodotto per procedere.", reply_markup=get_nav_markup("back_to_description"))
        bot.register_next_step_handler(msg, step_foto_process)
        return

    msg_wait = bot.send_message(message.chat.id, "🎨 **Grafica...**")
    
    if not os.path.exists("template.png"):
        bot.send_message(message.chat.id, "❌ Manca template.png")
        return

    file_info = bot.get_file(message.photo[-1].file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    
    try:
        img_prod = Image.open(io.BytesIO(downloaded_file)).convert("RGBA")
        template = Image.open("template.png").convert("RGBA")
        W, H = template.size 
        
        target_w = int(W * 0.45) 
        target_h = int(H * 0.70)
        ratio = min(target_w / img_prod.width, target_h / img_prod.height)
        new_w = int(img_prod.width * ratio)
        new_h = int(img_prod.height * ratio)
        img_prod = img_prod.resize((new_w, new_h), Image.Resampling.LANCZOS)
        
        mask = Image.new("L", (new_w, new_h), 0)
        draw_mask = ImageDraw.Draw(mask)
        draw_mask.rounded_rectangle((0, 0, new_w, new_h), radius=40, fill=255)
        img_rounded = Image.new("RGBA", (new_w, new_h), (0,0,0,0))
        img_rounded.paste(img_prod, (0,0), mask)
        img_prod = img_rounded

        pos_x = 80 
        pos_y = (H - new_h) // 2 + 30
        template.paste(img_prod, (pos_x, pos_y), img_prod)

        draw = ImageDraw.Draw(template)
        
        try:
            font_main = FONT_NAME if os.path.exists(FONT_NAME) else "arial.ttf"
            font_price = ImageFont.truetype(font_main, 90)
            font_disc = ImageFont.truetype(font_main, 85)
            font_small = ImageFont.truetype(font_main, 55)
            font_brand = ImageFont.truetype(font_main, 40)
        except:
            font_price = ImageFont.load_default()
            font_small = ImageFont.load_default()
            font_disc = ImageFont.load_default()
            font_brand = ImageFont.load_default()

        dati = user_data[message.chat.id]
        extras = dati.get('extras', {})
        COLORE_SCONTO = (204, 0, 0)
        COLORE_PREZZO = (228, 121, 17) 
        COLORE_VECCHIO = (120, 120, 120)
        COLORE_RISPARMIO = (34, 139, 34)
        MARGIN_RIGHT = 80 

        # 1. RISPARMIO IN ALTO
        risparmio_txt = dati['risparmio']
        if extras.get('coupon'):
             risparmio_txt = "COUPON DISPONIBILE"
             
        try:
            bbox_r = draw.textbbox((0, 0), risparmio_txt, font=font_brand)
            w_r = bbox_r[2] - bbox_r[0]
            x_r = (W - w_r) / 2
            draw.text((x_r, 75), risparmio_txt, font=font_brand, fill=COLORE_RISPARMIO)
        except: pass

        # 2. SCONTO
        testo_sconto = dati['sconto']
        
        if testo_sconto == "COUPON_MODE":
             testo_sconto = extras.get('coupon', 'NOVITÀ')
        elif testo_sconto == "N/D":
             testo_sconto = "" 
        
        if testo_sconto and testo_sconto != "AUMENTO": 
            disc_bbox = draw.textbbox((0, 0), testo_sconto, font=font_disc)
            disc_w = disc_bbox[2] - disc_bbox[0]
            draw.text((W - disc_w - MARGIN_RIGHT, 380), testo_sconto, font=font_disc, fill=COLORE_SCONTO)

        # --- 3. PREZZO NUOVO (AUTO-RESIZE) ---
        val_new_str = format_price_euro(dati['new']) + "€"
        dati['new_fmt'] = val_new_str 
        dati['new_fmt_save'] = format_price_for_excel(dati['new']) 

        current_font_size = 90
        foto_end_x = pos_x + new_w + 30
        text_end_x = W - MARGIN_RIGHT
        max_available_width = text_end_x - foto_end_x

        while True:
            font_price = ImageFont.truetype(font_main, current_font_size)
            price_bbox = draw.textbbox((0, 0), val_new_str, font=font_price)
            price_width = price_bbox[2] - price_bbox[0]
            
            if price_width < max_available_width:
                break
            
            current_font_size -= 5
            if current_font_size < 40: break 

        draw.text((text_end_x - price_width, 490), val_new_str, font=font_price, fill=COLORE_PREZZO)
        
        # 4. PREZZO VECCHIO (MATEMATICAMENTE CENTRATO)
        if dati['old'] > 0 and dati['old'] != dati['new']:
            val_old_str = format_price_euro(dati['old']) + "€"
            dati['old_fmt'] = val_old_str
            dati['old_fmt_save'] = format_price_for_excel(dati['old']) 
            
            old_bbox = draw.textbbox((0, 0), val_old_str, font=font_small)
            old_width = old_bbox[2] - old_bbox[0]
            old_height = old_bbox[3] - old_bbox[1]
            
            start_x_old = W - old_width - MARGIN_RIGHT
            start_y_old = 610
            
            draw.text((start_x_old, start_y_old), val_old_str, font=font_small, fill=COLORE_VECCHIO)
            
            line_y = start_y_old + (old_height / 2) + 12 
            draw.line((start_x_old, line_y, start_x_old + old_width, line_y), fill=COLORE_VECCHIO, width=8)
        else:
            dati['old_fmt_save'] = format_price_for_excel(dati['old'])
            dati['new_fmt_save'] = format_price_for_excel(dati['new'])


        bio = io.BytesIO()
        template = template.convert("RGB")
        template.save(bio, 'JPEG', quality=95)
        bio.seek(0)
        bot.delete_message(message.chat.id, msg_wait.message_id)
        
        # --- CAPTION ---
        txt_descrizione = ""
        if dati.get('descrizione'):
            txt_descrizione = f"{dati['descrizione']}\n\n" 
            
        txt_prime = "🚚 _Venduto e spedito da Amazon_\n✅ _Spedizione Prime_\n" if extras.get('prime') else "🚚 _Venduto e spedito da Amazon_\n"
        txt_lampo = "⏳ **OFFERTA A TEMPO**\n" if extras.get('lampo') else ""
        txt_choice = "⭐ **SCELTA AMAZON**\n" if extras.get('choice') else ""
        txt_coupon = f"✂️ **COUPON DISPONIBILE: {extras.get('coupon')}**\n" if extras.get('coupon') else ""
        txt_rapida = "🔥 **OFFERTA A VENDITA RAPIDA**\n" if extras.get('rapida') else "" 
        
        txt_reviews = ""
        if dati.get('voto'):
            txt_reviews = f"{dati['voto']}\n"
        
        titolo_prodotto = dati['titolo']
        emoji_prodotto = get_emoji_from_ia(titolo_prodotto)
        titolo_caption = f"{emoji_prodotto} {titolo_prodotto}"
        
        caption = (
            f"*{titolo_caption}*\n\n" 
            f"{txt_descrizione}" 
            f"ℹ️ _Dettagli su Amazon._\n\n" 
            f"{txt_lampo}"
            f"{txt_choice}"
            f"{txt_coupon}"
            f"{txt_rapida}" 
            f"{txt_prime}"
            f"➖➖➖➖➖➖➖➖➖➖\n"
            f"{txt_reviews}"
            f"\nℹ️ [Disclaimer & Info]({LINK_INFO_POST})"
        )
        
        # 💡 CORREZIONE: Controllo Lunghezza Caption (Limite 1024 caratteri)
        MAX_TELEGRAM_CAPTION = 1024
        if len(caption) > MAX_TELEGRAM_CAPTION:
            caption = caption[:MAX_TELEGRAM_CAPTION - 30] + "\n\n[...] *Caption Troncata*"
        
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("✅ Acquista Ora su Amazon ✅", url=dati['link']))
        markup.add(InlineKeyboardButton("🛒 Aggiungi al carrello", url=dati['cart_link']))
        share_text = urllib.parse.quote(f"Guarda: {dati['titolo']} a {dati['new_fmt']}! {dati['link']}")
        share_url = f"https://t.me/share/url?url={dati['link']}&text={share_text}"
        markup.add(InlineKeyboardButton("😏 Invita un amico", url=share_url))
        
        user_data[message.chat.id]['file_id'] = message.photo[-1].file_id 
        user_data[message.chat.id]['img_bytes'] = bio.getvalue()
        user_data[message.chat.id]['caption'] = caption
        user_data[message.chat.id]['markup'] = markup
        
        conf = InlineKeyboardMarkup()
        conf.add(InlineKeyboardButton("✅ PUBBLICA", callback_data="pubblica_si"))
        conf.add(InlineKeyboardButton("🔙 Cambia Foto/Extra", callback_data="back_to_description")) 
        conf.add(InlineKeyboardButton("❌ ANNULLA", callback_data="reset_all"))
        
        bot.send_photo(message.chat.id, bio, caption=caption, reply_markup=markup, parse_mode='Markdown')
        bot.send_message(message.chat.id, "Anteprima perfetta! Sei pronto per pubblicare sul canale?", reply_markup=conf)

    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Errore: {e}", parse_mode='Markdown')
        handle_critical_error(e) # Invio notifica per errore grafico

# --- NUOVO FLUSSO DI LAVORO CON AUTOMAZIONE (Invariato) ---

# STEP 1: step_link (MODIFICATA)
def step_link(message):
    chat_id = message.chat.id
    if message.text and message.text.startswith('/'): return
    
    if chat_id not in user_data:
        user_data[chat_id] = {'extras': {'prime': False, 'lampo': False, 'choice': False, 'coupon': None, 'rapida': False}}
        
    link_raw = message.text
    
    if not link_raw or not "http" in link_raw:
        bot.send_message(chat_id, "❌ Per favore, incolla un link *valido* che contenga 'http'.", parse_mode='Markdown')
        ask_link(chat_id) 
        return

    asin = "NO_ASIN"
    link_aff = link_raw 
    cart_link = link_raw
    
    try:
        asin_match = re.search(r'(B0[A-Z0-9]{8})', link_raw)
        if asin_match:
            asin = asin_match.group(1)
            link_aff = f"https://www.amazon.it/dp/{asin}?tag={AMAZON_TAG}"
            cart_link = f"https://www.amazon.it/gp/aws/cart/add.html?ASIN.1={asin}&Quantity.1=1&tag={AMAZON_TAG}"
            if is_gia_pubblicato(asin):
                bot.send_message(chat_id, f"⚠️ ATTENZIONE: Già pubblicato oggi!", parse_mode='Markdown')
        else:
            sep = "&" if "?" in link_raw else "?"
            link_aff = f"{link_raw}{sep}tag={AMAZON_TAG}"
            cart_link = link_aff 
        
        user_data[chat_id]['link'] = link_aff
        user_data[chat_id]['cart_link'] = cart_link
        user_data[chat_id]['asin'] = asin

        # NUOVA LOGICA DI ESTRAZIONE DATI
        msg_wait = bot.send_message(chat_id, "🔎 **Sto estraendo titolo e prezzo dalla pagina Amazon...**", parse_mode='Markdown')
        
        scraped_data = get_product_data(link_aff)
        bot.delete_message(chat_id, msg_wait.message_id)

        if scraped_data:
            user_data[chat_id]['titolo'] = scraped_data['titolo_estratto']
            user_data[chat_id]['new'] = scraped_data['prezzo_nuovo_estratto']
            
            prezzo_fmt = format_price_euro(scraped_data['prezzo_nuovo_estratto']) + "€"
            
            ask_confirm_data(chat_id, scraped_data['titolo_estratto'], prezzo_fmt)
            
        else:
            # Fallback al flusso manuale
            bot.send_message(chat_id, "⚠️ **Estrazione fallita.** Procediamo con l'inserimento manuale (Chiedo il titolo prima).", parse_mode='Markdown')
            
            # 💡 CORREZIONE FLOW: Ripristino a ask_title(chat_id) che avvia la sequenza manuale
            ask_title(chat_id) 

    except Exception as e:
        bot.send_message(chat_id, f"❌ Errore durante l'analisi del link: {e}. Riprova.", parse_mode='Markdown')
        handle_critical_error(e)
        ask_link(chat_id)


# STEP 2: Conferma dei dati estratti e richiesta del prezzo vecchio
def ask_confirm_data(chat_id, titolo_estratto, prezzo_fmt):
    """Chiede all'utente di confermare i dati estratti e di inserire il prezzo vecchio."""
    bot.clear_step_handler_by_chat_id(chat_id)
    
    messaggio = (
        "✅ **Dati Estratti (Conferma):**\n\n"
        f"**Titolo:** `{titolo_estratto}`\n"
        f"**Prezzo NUOVO:** `{prezzo_fmt}`\n\n"
        "2️⃣ **Inserire il PREZZO VECCHIO** (es: 1499.00 o 1.499,00) "
        "oppure usa i pulsanti per correggere."
    )
    
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("✍️ Correggi Titolo", callback_data="correct_title_auto"))
    markup.add(InlineKeyboardButton("✍️ Correggi Nuovo Prezzo", callback_data="correct_new_price_auto"))
    markup.add(InlineKeyboardButton("❌ ANNULLA POST", callback_data="reset_all"))
    
    msg = bot.send_message(chat_id, messaggio, reply_markup=markup, parse_mode='Markdown')
    bot.register_next_step_handler(msg, step_old_price_or_correction)


# STEP HANDLER: Gestisce l'input del prezzo vecchio o i comandi di correzione
def step_old_price_or_correction(message):
    chat_id = message.chat.id
    text = message.text

    if text and text.startswith('/'): 
        return

    if message.content_type == 'text':
        pulito = clean_input_price(text)
        val = clean_price_calc(pulito) 
        
        if val > 0.0:
            # L'utente ha inserito un prezzo valido, lo consideriamo il prezzo vecchio.
            user_data[chat_id]['old'] = val
            
            # Calcola sconto con i 3 prezzi (vecchio, nuovo estratto, nuovo pulito)
            step_update_discount(chat_id)
            
            # Passiamo allo step delle recensioni (saltando la richiesta manuale di titolo e nuovo prezzo)
            ask_reviews(chat_id)
            return

    # Se non è un prezzo valido, è probabile che l'utente voglia correggere qualcosa. 
    bot.send_message(chat_id, "❌ Input non valido. Inserire solo il PREZZO VECCHIO (es. 120,50) o usare i pulsanti.", parse_mode='Markdown')
    # Ri-visualizziamo il messaggio di conferma per non perdere il contesto
    prezzo_fmt = format_price_euro(user_data[chat_id]['new']) + "€"
    ask_confirm_data(chat_id, user_data[chat_id]['titolo'], prezzo_fmt)


# NUOVE FUNZIONI DI CORREZIONE
def step_correct_title(message):
    if message.text and message.text.startswith('/'): return
    chat_id = message.chat.id
    titolo = rewrite_with_ai(message.text)
    user_data[chat_id]['titolo'] = titolo
    
    prezzo_fmt = format_price_euro(user_data[chat_id]['new']) + "€"
    bot.send_message(chat_id, f"✅ Titolo aggiornato: ✨ {titolo}", parse_mode='Markdown')
    ask_confirm_data(chat_id, user_data[chat_id]['titolo'], prezzo_fmt)

def step_correct_new_price(message):
    if message.text and message.text.startswith('/'): return
    chat_id = message.chat.id
    pulito = clean_input_price(message.text)
    val = clean_price_calc(pulito) 
    
    if val > 0.0:
        user_data[chat_id]['new'] = val
        prezzo_fmt = format_price_euro(val) + "€"
        bot.send_message(chat_id, f"✅ Nuovo Prezzo aggiornato: `{prezzo_fmt}`", parse_mode='Markdown')
        ask_confirm_data(chat_id, user_data[chat_id]['titolo'], prezzo_fmt)
    else:
        bot.send_message(chat_id, "❌ Prezzo Nuovo non valido. Riprova.", parse_mode='Markdown')
        msg = bot.send_message(chat_id, "✍️ **Inserire il NUOVO Prezzo Corretto:**", reply_markup=get_nav_markup("back_to_link"), parse_mode='Markdown')
        bot.register_next_step_handler(msg, step_correct_new_price)


# step_update_discount (Calcola lo sconto)
def step_update_discount(chat_id):
    """Calcola lo sconto e il risparmio una volta che tutti i prezzi sono disponibili."""
    try:
        old_val = user_data[chat_id]['old']
        new_val = user_data[chat_id]['new'] 
        extras = user_data.get(chat_id, {}).get('extras', {})
        
        st = "N/D" 
        risp_str = "Nessun Risparmio"
        
        if old_val > new_val:
            perc = int(100 - (new_val / old_val * 100))
            st = f"-{perc}%"
            risp = old_val - new_val
            risp_str = f"RISPARMI: {format_price_euro(risp)}€" 
        
        elif old_val == new_val or old_val == 0:
            if extras.get('coupon'):
                st = "COUPON_MODE" 
                risp_str = "COUPON DISPONIBILE"
            else:
                st = "N/D" 
                risp_str = "Nessun Risparmio"
        
        elif old_val < new_val:
            st = "AUMENTO"
            risp_str = "Prezzo Aumentato"
            
        user_data[chat_id]['sconto'] = st
        user_data[chat_id]['risparmio'] = risp_str
        
    except Exception as e: 
         print(f"Errore calcolo sconto in step_update_discount: {e}")
         user_data[chat_id]['sconto'] = "N/D"
         user_data[chat_id]['risparmio'] = "Errore Calcolo"

# --- MAIN LOOP CORRETTO ---
if __name__ == '__main__':
    inizializza_db()
    
    # Loop infinito con gestione degli errori critici
    while True:
        try:
            # Controllo token prima di avviare il polling
            if not API_TOKEN:
                print("Impossibile avviare il bot. Rivedere la configurazione TELEGRAM_BOT_TOKEN nel file .env.")
                time.sleep(60) # Attendi prima di un eventuale riavvio automatico del container
                continue

            bot.infinity_polling(timeout=10, long_polling_timeout=5)
        except Exception as e:
            handle_critical_error(e) 
            time.sleep(5)