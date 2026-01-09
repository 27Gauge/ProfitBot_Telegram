import os
import time
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
from dotenv import load_dotenv
import telebot
import random
import traceback

# --- CONFIGURAZIONE ---
load_dotenv()
# Assicurati che TELEGRAM_BOT_TOKEN e FABRIZIO_CHAT_ID siano nel tuo file .env
API_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN') 
FABRIZIO_CHAT_ID = os.getenv('FABRIZIO_CHAT_ID') 
DB_FILE = "Registro_Vendite.xlsx"
WATCHLIST_FILE = "watchlist.txt"
AMAZON_TAG = 'radartest-21' 

bot = telebot.TeleBot(API_TOKEN)
print("✅ Cruiser Bot Avviato. Monitoraggio DB.")


# --- UTILITY E DB ---

def handle_critical_error(e, context="CRUISER"):
    error_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    error_traceback = traceback.format_exc()
    message = (
        f"🚨 **ERRORE CRITICO in {context}!** 🚨\n"
        f"⏰ Data/Ora: `{error_time}`\n"
        f"💬 Messaggio: `{str(e)}`\n"
        f"--- TRACEBACK ---\n"
        f"```\n"
        f"{error_traceback[:500]}..." 
        f"```"
    )
    if FABRIZIO_CHAT_ID:
        try:
            bot.send_message(FABRIZIO_CHAT_ID, message, parse_mode='Markdown')
        except Exception as notify_e:
            print(f"Impossibile inviare la notifica: {notify_e}")
    print(message) 

def clean_price_calc(price_str):
    if not isinstance(price_str, str): return 0.0
    price_str = price_str.replace('€', '').replace('$', '').strip()
    # Logica per gestire il separatore decimale italiano
    if price_str.count(',') == 1 and price_str.count('.') <= 1 and price_str.rfind(',') > price_str.rfind('.'):
        price_str = price_str.replace('.', '')
        price_str = price_str.replace(',', '.')
    else:
        price_str = price_str.replace(',', '') 
        
    try:
        return float(price_str)
    except ValueError:
        return 0.0 

def format_price_for_excel(val_float):
    return "{:.2f}".format(val_float).replace('.', ',')

def get_last_price_from_db(asin):
    if not os.path.exists(DB_FILE): return 0.0, 0.0
    try:
        wb = load_workbook(DB_FILE)
        ws = wb.active
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=True))):
            db_asin = row[2]
            if db_asin == asin:
                # [5] Nuovo Prezzo è il più recente registrato
                return clean_price_calc(row[5]), clean_price_calc(row[4]) 
        return 0.0, 0.0
    except Exception as e:
        handle_critical_error(e, "DB_READ")
        return 0.0, 0.0

def salva_in_excel(dati, tipo_pubblicazione="CRUISER_MONITOR"):
    adesso = datetime.now()
    try:
        if not os.path.exists(DB_FILE):
             # Crea un nuovo file se non esiste
             wb = Workbook()
             ws = wb.active
             ws.title = "Registro Vendite"
             ws.append(["DATA", "ORA", "ASIN", "TITOLO", "PREZZO_VECCHIO", "PREZZO_NUOVO", "LINK_AFFILIATO", "FILE_ID", "TIPO_PUBBLICAZIONE"])
        else:
            wb = load_workbook(DB_FILE)
            ws = wb.active

        # Assicurati che la colonna TIPO_PUBBLICAZIONE esista
        if ws.cell(row=1, column=9).value != "TIPO_PUBBLICAZIONE":
            ws.cell(row=1, column=9, value="TIPO_PUBBLICAZIONE")
        
        ws.append([
            adesso.strftime("%d/%m/%Y"), 
            adesso.strftime("%H:%M:%S"), 
            dati.get('asin', 'N/A'), 
            dati.get('titolo', 'Nessun Titolo'), 
            dati.get('old_price_db_save', '0,00'), 
            dati.get('new_price_scraped_save', '0,00'), 
            dati.get('link_aff', ''),
            'CRUISER_N/A', 
            tipo_pubblicazione
        ])
        wb.save(DB_FILE)
        return True
    except Exception as e:
        handle_critical_error(e, "DB_WRITE")
        return False

# --- FUNZIONE WEB SCRAPING GENTILE (VERS. DEFINITIVA) ---
def get_product_data_gentle(url):
    """Estrae titolo e prezzo corrente da un link Amazon con logica di fallback e OOS."""
    
    asin_match = re.search(r'(B0[A-Z0-9]{8})', url)
    if not asin_match:
        print(f"⚠️ ASIN non trovato in {url}. Salto.")
        return None, None
        
    asin = asin_match.group(1)
    url_aff = f"https://www.amazon.it/dp/{asin}?tag={AMAZON_TAG}"
    
    headers = {
        'User-Agent': random.choice([
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Safari/605.1.15',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36'
        ]),
        'Accept-Language': 'it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7',
    }
    
    try:
        # ⚠️ SCRAPING GENTILE: Ritardo casuale
        sleep_time = random.uniform(5, 15) 
        time.sleep(sleep_time)
        
        response = requests.get(url_aff, headers=headers, timeout=15)
        
        if response.status_code != 200:
            print(f"❌ Scraping fallito per {asin}: Status {response.status_code}. Riprova.")
            return None, None

        if "captcha" in response.text.lower() or response.status_code == 503:
             print("❌ Scraping fallito: CAPTCHA o blocco Amazon (Status 503).")
             # Se necessario, qui si può inserire la logica di pausa lunga (es. 1 ora)
             return None, None 

        soup = BeautifulSoup(response.content, 'html.parser')
        
        # 1. ESTREZIONE TITOLO
        title_tag = soup.find('span', id='productTitle')
        title = title_tag.text.strip() if title_tag else "Titolo non trovato"

        # 2. ESTREZIONE PREZZO (LOGICA AGGIORNATA)
        price = None
        
        # Cerca 1: Prezzo principale (a-price con a-offscreen)
        price_tag_offscreen = soup.find('span', class_='a-price')
        if price_tag_offscreen:
            price_text = price_tag_offscreen.find('span', class_='a-offscreen')
            if price_text:
                price = price_text.text.strip()
        
        # Fallback 1: Cerca il prezzo per offerte lampo/prezzo barrato
        if not price:
            price_tag_main = soup.find('span', id='priceblock_ourprice') or soup.find('span', id='priceblock_dealprice')
            if price_tag_main:
                price = price_tag_main.text.strip()
        
        # Fallback 2: Cerca il prezzo per le variazioni/core price feature
        if not price:
             price_tag_variation = soup.find('div', id='corePriceDisplay_desktop_feature_div')
             if price_tag_variation:
                 price_text = price_tag_variation.find('span', class_='a-offscreen')
                 if price_text:
                      price = price_text.text.strip()
                      
        # ⭐⭐ NUOVO FALLBACK 3: CERCA GENERICAMENTE a-offscreen (Ultima Risorsa) ⭐⭐
        if not price:
            all_offscreen = soup.find_all('span', class_='a-offscreen')
            # Prendiamo il primo a-offscreen che contiene la valuta, spesso è il prezzo principale.
            for tag in all_offscreen:
                price_text = tag.text.strip()
                if '€' in price_text or '$' in price_text:
                    price = price_text
                    break
        
        # 3. CONTROLLO OOS (Esaurimento Scorte)
        oos_tag = soup.find('div', id='availability')
        if oos_tag and ("non disponibile" in oos_tag.text.lower() or "currently unavailable" in oos_tag.text.lower()):
            print(f"⚠️ Prodotto {asin} Esaurito (OOS). Prezzo impostato a 0.0")
            cleaned_price = 0.0
        else:
            cleaned_price = clean_price_calc(price)
            
        if cleaned_price > 0.0:
            return asin, {
                'titolo': title,
                'prezzo_nuovo': cleaned_price,
                'link_aff': url_aff
            }
        
        print(f"⚠️ Dati incompleti per {asin}. Prezzo pulito: {cleaned_price}. Titolo: {title}")
        return None, None

    except requests.exceptions.Timeout:
        print(f"❌ Timeout Scraping per {url_aff}.")
        return None, None
    except Exception as e:
        handle_critical_error(e, f"SCRAPER: {url_aff}")
        return None, None

# --- FUNZIONE PRINCIPALE DEL CRUISER ---
def run_cruiser():
    
    try:
        with open(WATCHLIST_FILE, 'r') as f:
            urls = [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"❌ File {WATCHLIST_FILE} non trovato. Impossibile avviare il cruiser.")
        return

    print(f"⏳ Inizio ciclo di monitoraggio su {len(urls)} prodotti...")

    for url in urls:
        asin, data_scraped = get_product_data_gentle(url)
        
        if not asin:
            continue

        current_price = data_scraped['prezzo_nuovo']
        
        # Se lo scraper ha trovato il prodotto ma il prezzo è 0 (OOS)
        if current_price == 0.0:
            continue

        last_price_db, _ = get_last_price_from_db(asin)
        
        if last_price_db == 0.0:
             # Primo monitoraggio: usiamo il prezzo corrente come 'vecchio'
             last_price_db = current_price
             
        # --- LOGICA DI RILEVAMENTO RIBASSO ---
        
        # Se il prezzo è sceso di almeno il 10% O è sceso sotto l'ultimo prezzo noto
        is_significant_drop = last_price_db > 0 and (current_price / last_price_db) < 0.90
        is_lower_than_last = current_price < last_price_db
        
        if is_significant_drop or is_lower_than_last:
            
            difference = last_price_db - current_price
            perc_drop = int(100 - (current_price / last_price_db * 100)) if last_price_db > 0 else 0
            
            dati_salvataggio = {
                'asin': asin,
                'titolo': data_scraped['titolo'],
                'old_price_db_save': format_price_for_excel(last_price_db),
                'new_price_scraped_save': format_price_for_excel(current_price),
                'link_aff': data_scraped['link_aff']
            }
            
            if salva_in_excel(dati_salvataggio, "CRUISER_RIBASSO"):
                
                message_notify = (
                    f"🔔 **🚨 NUOVO RIBASSO AUTOMATICO RILEVATO!** 🚨\n\n"
                    f"**Prodotto:** {data_scraped['titolo']}\n"
                    f"**Prezzo Corrente:** `{format_price_for_excel(current_price)}€`\n"
                    f"**Prezzo Precedente (DB):** `{format_price_for_excel(last_price_db)}€`\n"
                    f"📉 **Sconto Rilevato:** -{perc_drop}% ({format_price_for_excel(difference)}€)\n\n"
                    f"👉 **PRONTO PER LA PUBBLICAZIONE:** Vai al menu `Inizia Nuovo Post` in ProfitBot.\n"
                    f"   Il prodotto è l'ultimo salvato nel tuo database."
                )
                
                try:
                    bot.send_message(FABRIZIO_CHAT_ID, message_notify, parse_mode='Markdown')
                except Exception as e:
                    handle_critical_error(e, "TELEGRAM_NOTIFY")
                
            else:
                 print(f"❌ Errore salvataggio DB per {asin}.")

        elif current_price > last_price_db and last_price_db > 0:
            print(f"📈 Prezzo aumentato per {asin}. Non registrato.")
            pass
            
        else:
             print(f"✅ Prodotto {asin} monitorato. Nessun cambiamento significativo.")

    print("--- Ciclo completato. Riavvio in 5 minuti. ---")
    time.sleep(300) 

# --- MAIN LOOP PER IL CRUISER ---
if __name__ == '__main__':
    if not API_TOKEN or not FABRIZIO_CHAT_ID:
        print("❌ ERRORE: TELEGRAM_BOT_TOKEN o FABRIZIO_CHAT_ID mancante in .env. Impossibile avviare.")
    else:
        while True:
            try:
                # Controlla se Amazon ha bloccato prima di eseguire
                if os.path.exists('PAUSE_CRUISER.flag'):
                    print("Cruiser in pausa. Riprovo tra 1 minuto.")
                    time.sleep(60)
                    continue

                run_cruiser()
                
            except Exception as e:
                # Se ricevi un blocco da Amazon (es. 503) e viene sollevata un'eccezione
                if "BLOCCO AMAZON" in str(e) or "503" in str(e):
                    # Crea un flag per mettere in pausa
                    with open('PAUSE_CRUISER.flag', 'w') as f:
                        f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    bot.send_message(FABRIZIO_CHAT_ID, "⚠️ **BLOCCO AMAZON RILEVATO.** Il cruiser è in pausa per 1 ora.", parse_mode='Markdown')
                    time.sleep(3600)
                    # Qui dovresti implementare la logica per cambiare IP (se usi un proxy avanzato)
                    os.remove('PAUSE_CRUISER.flag') # Rimuovi il flag e riprova
                    
                else:
                    handle_critical_error(e, "MAIN_LOOP")
                    print("❌ Errore nel loop principale. Riavvio in 30 secondi.")
                    time.sleep(30)